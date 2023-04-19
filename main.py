from typing import Optional
from hdwallet import BIP44HDWallet
from hdwallet.cryptocurrencies import EthereumMainnet
from hdwallet.utils import generate_mnemonic
import openpyxl

# Подготовка таблицы для сохранения
book = openpyxl.Workbook()
sheet = book.active
sheet['A1'] = "Mnemonic"
sheet['B1'] = "Address"
sheet['C1'] = "Private key"

# Выбор кол-ва кошельков для генерации
#N = int(input("Enter number of wallets: "))

#for i in range(N):
for i in range(100):
    MNEMONIC: str = generate_mnemonic(language="english", strength=128)
    PASSPHRASE: Optional[str] = None  # "meherett"
    bip44_hdwallet: BIP44HDWallet = BIP44HDWallet(cryptocurrency=EthereumMainnet)
    bip44_hdwallet.from_mnemonic(mnemonic=MNEMONIC, language="english", passphrase=PASSPHRASE)
    sheet.cell(row=i+2,column=1).value=bip44_hdwallet.mnemonic()
    sheet.cell(row=i+2,column=2).value=bip44_hdwallet.address()
    sheet.cell(row=i+2, column=3).value=bip44_hdwallet.private_key()
    bip44_hdwallet.clean_derivation()

book.save("wallets.xlsx")
book.close()
